import os
import io
import csv
import re
import copy
import logging
import tempfile
from collections import defaultdict
from typing import Any, Dict, List, Optional, Tuple

import vobject
import phonenumbers
from phonenumbers import geocoder, carrier
import openpyxl
from openpyxl import Workbook

from telegram import (
    Update,
    InputFile,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    KeyboardButton,
)
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ConversationHandler,
    ContextTypes,
    filters,
)
from telegram.constants import ParseMode

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

BOT_TOKEN = "7727685861:AAGrMehK3GiBK2cR_WO1z4IrwJqJcFUtxfs"

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# CONVERSATION STATES
# ─────────────────────────────────────────────────────────────────────────────

# Main menu
MAIN_MENU = 0

# File Analysis
ANALYSIS_UPLOAD = 10

# File Converter
CONVERT_UPLOAD = 20
CONVERT_FORMAT = 21

# Quick VCF
QVCF_FILENAME = 30
QVCF_CNAME = 31
QVCF_CNUMBER = 32
QVCF_MORE = 33

# VCF Maker
MAKER_FILENAME = 40
MAKER_CNAME = 41
MAKER_PER_FILE = 42
MAKER_NUM_START = 43
MAKER_FILE_START = 44
MAKER_COUNTRY = 45
MAKER_GROUP = 46
MAKER_CONFIRM = 47
MAKER_EDIT_FIELD = 48
MAKER_EDIT_WHICH = 49
MAKER_UPLOAD = 50
MAKER_OUTPUT_MODE = 51

# Split File
SPLIT_UPLOAD = 60
SPLIT_COUNT = 61

# Merge Files
MERGE_UPLOAD = 70

# File Editor
EDITOR_UPLOAD = 80
EDITOR_MENU = 81
EDITOR_REMOVE = 82
EDITOR_EDIT_SEL = 83
EDITOR_EDIT_NAME = 84
EDITOR_EDIT_NUM = 85
EDITOR_ADD_NAME = 86
EDITOR_ADD_NUM = 87

# List Maker
LIST_NAME = 90
LIST_UPLOAD = 91

# Rename File
RENFILE_UPLOAD = 100
RENFILE_NAME = 101

# Rename Contact
RENCONT_UPLOAD = 110
RENCONT_NAME = 111

# Settings
SETTINGS_MENU = 120
SETTINGS_EDIT = 121

# Reset
RESET_CONFIRM = 130

# ─────────────────────────────────────────────────────────────────────────────
# BUTTON TEXT CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

# Main menu buttons
BTN_ANALYSIS = "📊 File Analysis"
BTN_CONVERTER = "🔄 File Converter"
BTN_QVCF = "⚡ Quick VCF"
BTN_MAKER = "🏭 VCF Maker"
BTN_SPLIT = "✂️ Split File"
BTN_MERGE = "🔗 Merge Files"
BTN_EDITOR = "✏️ File Editor"
BTN_LISTMAKER = "📋 List Maker"
BTN_RENFILE = "📝 Rename File"
BTN_RENCONT = "📛 Rename Contact"
BTN_SETTINGS = "⚙️ Settings"
BTN_RESET = "🔄 Reset"
BTN_HELP = "❓ Help"

# Common action buttons
BTN_CANCEL = "❌ Cancel"
BTN_BACK = "🔙 Back"
BTN_FINISH = "✅ Finish"
BTN_YES = "✅ Yes"
BTN_NO = "❌ No"
BTN_ADD_MORE = "➕ Add More"
BTN_UPLOAD_MORE = "📎 Upload More"
BTN_FINISH_UPLOAD = "✅ Finish Uploading"
BTN_FINISH_MERGING = "✅ Finish Merging"
BTN_GENERATE = "🏁 Finish & Generate"

# Format buttons
BTN_FMT_VCF = "📇 VCF"
BTN_FMT_TXT = "📄 TXT"
BTN_FMT_EXCEL = "📊 Excel (XLSX)"
BTN_FMT_CSV = "📑 CSV"

# Yes/No reset
BTN_YES_RESET = "✅ Yes, Reset All"
BTN_NO_CANCEL = "❌ No, Cancel"

# Output mode
BTN_SINGLE = "📦 Single Merged File"
BTN_SEPARATE = "📂 Separate Files"

# Finish generating
BTN_FINISH_GENERATE = "🏁 Generate List"

# Editor actions
BTN_REMOVE = "➖ Remove Contact"
BTN_EDIT = "✏️ Edit Contact"
BTN_ADD = "➕ Add Contact"
BTN_DOWNLOAD = "💾 Finish & Download"

# Settings edit buttons
BTN_EDIT_1 = "1️⃣ File Name"
BTN_EDIT_2 = "2️⃣ Contact Name"
BTN_EDIT_3 = "3️⃣ Per File"
BTN_EDIT_4 = "4️⃣ Num Start"
BTN_EDIT_5 = "5️⃣ File Start"
BTN_EDIT_6 = "6️⃣ Country Code"
BTN_EDIT_7 = "7️⃣ Group Name"

# Maker confirm edit buttons
BTN_MAKER_CONFIRM = "✅ Confirm & Upload Files"

# ─────────────────────────────────────────────────────────────────────────────
# DEFAULT SETTINGS (per user, in-memory)
# ─────────────────────────────────────────────────────────────────────────────

DEFAULT_SETTINGS: Dict[str, Any] = {
    "vcf_filename": "Contacts",
    "contact_name": "Contact",
    "per_file": 1000,
    "num_start": 1,
    "file_start": 1,
    "country_code": "+1",
    "group_name": "Group",
}

USER_SETTINGS: Dict[int, Dict[str, Any]] = {}
USER_DATA: Dict[int, Dict[str, Any]] = {}


def get_settings(uid: int) -> Dict[str, Any]:
    if uid not in USER_SETTINGS:
        USER_SETTINGS[uid] = copy.deepcopy(DEFAULT_SETTINGS)
    return USER_SETTINGS[uid]


def get_udata(uid: int) -> Dict[str, Any]:
    if uid not in USER_DATA:
        USER_DATA[uid] = {}
    return USER_DATA[uid]


def reset_udata(uid: int) -> None:
    USER_DATA[uid] = {}


# ─────────────────────────────────────────────────────────────────────────────
# FLAG EMOJI HELPER
# ─────────────────────────────────────────────────────────────────────────────

COUNTRY_FLAGS: Dict[str, str] = {
    "US": "🇺🇸", "GB": "🇬🇧", "IN": "🇮🇳", "PK": "🇵🇰", "BD": "🇧🇩",
    "NG": "🇳🇬", "PH": "🇵🇭", "ID": "🇮🇩", "BR": "🇧🇷", "MX": "🇲🇽",
    "DE": "🇩🇪", "FR": "🇫🇷", "IT": "🇮🇹", "ES": "🇪🇸", "TR": "🇹🇷",
    "RU": "🇷🇺", "CN": "🇨🇳", "JP": "🇯🇵", "KR": "🇰🇷", "SA": "🇸🇦",
    "AE": "🇦🇪", "EG": "🇪🇬", "ZA": "🇿🇦", "KE": "🇰🇪", "GH": "🇬🇭",
    "TZ": "🇹🇿", "ET": "🇪🇹", "MA": "🇲🇦", "DZ": "🇩🇿", "TN": "🇹🇳",
    "CA": "🇨🇦", "AU": "🇦🇺", "NZ": "🇳🇿", "AR": "🇦🇷", "CO": "🇨🇴",
    "VE": "🇻🇪", "CL": "🇨🇱", "PE": "🇵🇪", "PL": "🇵🇱", "NL": "🇳🇱",
    "BE": "🇧🇪", "SE": "🇸🇪", "NO": "🇳🇴", "DK": "🇩🇰", "FI": "🇫🇮",
    "PT": "🇵🇹", "GR": "🇬🇷", "UA": "🇺🇦", "RO": "🇷🇴", "HU": "🇭🇺",
    "CZ": "🇨🇿", "AT": "🇦🇹", "CH": "🇨🇭", "TH": "🇹🇭", "VN": "🇻🇳",
    "MY": "🇲🇾", "SG": "🇸🇬", "NP": "🇳🇵", "LK": "🇱🇰", "MM": "🇲🇲",
    "KH": "🇰🇭", "AF": "🇦🇫", "IR": "🇮🇷", "IQ": "🇮🇶", "SY": "🇸🇾",
    "JO": "🇯🇴", "LB": "🇱🇧", "KW": "🇰🇼", "QA": "🇶🇦", "BH": "🇧🇭",
    "OM": "🇴🇲", "YE": "🇾🇪", "UZ": "🇺🇿", "KZ": "🇰🇿", "AZ": "🇦🇿",
    "GE": "🇬🇪", "AM": "🇦🇲", "IL": "🇮🇱", "BY": "🇧🇾",
}


def get_flag(cc: str) -> str:
    return COUNTRY_FLAGS.get(cc.upper(), "🌍")


# ─────────────────────────────────────────────────────────────────────────────
# PHONE NUMBER UTILITIES
# ─────────────────────────────────────────────────────────────────────────────

def clean_number(raw: str) -> str:
    raw = raw.strip()
    digits = re.sub(r"[^\d+]", "", raw)
    return digits


def parse_phone(raw: str) -> Optional[phonenumbers.PhoneNumber]:
    try:
        cleaned = clean_number(raw)
        if not cleaned:
            return None
        if cleaned.startswith("+"):
            return phonenumbers.parse(cleaned, None)
        else:
            return phonenumbers.parse("+" + cleaned, None)
    except Exception:
        return None


def is_valid_number(raw: str) -> bool:
    pn = parse_phone(raw)
    if pn is None:
        return False
    return phonenumbers.is_valid_number(pn)


def get_country_code_str(raw: str) -> Optional[str]:
    pn = parse_phone(raw)
    if pn is None:
        return None
    return phonenumbers.region_code_for_number(pn)


def format_e164(raw: str) -> str:
    pn = parse_phone(raw)
    if pn is None:
        return clean_number(raw)
    return phonenumbers.format_number(pn, phonenumbers.PhoneNumberFormat.E164)


# ─────────────────────────────────────────────────────────────────────────────
# FILE PARSING UTILITIES
# ─────────────────────────────────────────────────────────────────────────────

def parse_vcf_bytes(data: bytes) -> List[Tuple[str, str]]:
    contacts = []
    try:
        text = data.decode("utf-8", errors="ignore")
        for vcard in vobject.readComponents(text):
            name = ""
            number = ""
            try:
                name = str(vcard.fn.value).strip()
            except Exception:
                pass
            try:
                tel = vcard.tel.value
                number = clean_number(str(tel))
            except Exception:
                pass
            if number:
                contacts.append((name, number))
    except Exception as e:
        logger.error(f"VCF parse error: {e}")
    return contacts


def parse_txt_bytes(data: bytes) -> List[Tuple[str, str]]:
    contacts = []
    text = data.decode("utf-8", errors="ignore")
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        parts = line.split(None, 1)
        if len(parts) == 2 and not parts[0].lstrip("+").isdigit():
            contacts.append((parts[0], clean_number(parts[1])))
        else:
            contacts.append(("", clean_number(line)))
    return contacts


def parse_csv_bytes(data: bytes) -> List[Tuple[str, str]]:
    contacts = []
    text = data.decode("utf-8", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return contacts
    start = 0
    header = [h.lower().strip() for h in rows[0]]
    name_col, num_col = 0, 1
    for i, h in enumerate(header):
        if "name" in h:
            name_col = i
        if "phone" in h or "number" in h or "mobile" in h or "tel" in h:
            num_col = i
    if any(("name" in h or "phone" in h or "number" in h) for h in header):
        start = 1
    for row in rows[start:]:
        if not row:
            continue
        if len(row) == 1:
            num = clean_number(row[0])
            contacts.append(("", num))
        else:
            try:
                name = row[name_col].strip() if name_col < len(row) else ""
                num = clean_number(row[num_col]) if num_col < len(row) else ""
            except Exception:
                name, num = "", ""
            contacts.append((name, num))
    return contacts


def parse_xlsx_bytes(data: bytes) -> List[Tuple[str, str]]:
    contacts = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return contacts
        start = 0
        name_col, num_col = 0, 1
        if rows:
            first = [str(c).lower().strip() if c else "" for c in rows[0]]
            for i, h in enumerate(first):
                if "name" in h:
                    name_col = i
                if "phone" in h or "number" in h or "mobile" in h or "tel" in h:
                    num_col = i
            if any("name" in h or "phone" in h or "number" in h for h in first):
                start = 1
        for row in rows[start:]:
            if not row:
                continue
            name = str(row[name_col]).strip() if name_col < len(row) and row[name_col] else ""
            num = clean_number(str(row[num_col])) if num_col < len(row) and row[num_col] else ""
            if num:
                contacts.append((name, num))
    except Exception as e:
        logger.error(f"XLSX parse error: {e}")
    return contacts


def detect_format(filename: str) -> str:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    return ext


def parse_file_bytes(data: bytes, filename: str) -> List[Tuple[str, str]]:
    fmt = detect_format(filename)
    if fmt == "vcf":
        return parse_vcf_bytes(data)
    elif fmt == "csv":
        return parse_csv_bytes(data)
    elif fmt in ("xlsx", "xls"):
        return parse_xlsx_bytes(data)
    else:
        return parse_txt_bytes(data)


# ─────────────────────────────────────────────────────────────────────────────
# FILE GENERATION UTILITIES
# ─────────────────────────────────────────────────────────────────────────────

def build_vcf_bytes(contacts: List[Tuple[str, str]]) -> bytes:
    lines = []
    for name, number in contacts:
        lines.append("BEGIN:VCARD")
        lines.append("VERSION:3.0")
        lines.append(f"FN:{name}")
        lines.append(f"TEL;TYPE=CELL:{number}")
        lines.append("END:VCARD")
    return "\n".join(lines).encode("utf-8")


def build_txt_bytes(contacts: List[Tuple[str, str]]) -> bytes:
    lines = []
    for name, number in contacts:
        if name:
            lines.append(f"{name} {number}")
        else:
            lines.append(number)
    return "\n".join(lines).encode("utf-8")


def build_csv_bytes(contacts: List[Tuple[str, str]]) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Name", "Phone"])
    for name, number in contacts:
        writer.writerow([name, number])
    return output.getvalue().encode("utf-8")


def build_xlsx_bytes(contacts: List[Tuple[str, str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    ws.append(["Name", "Phone"])
    for name, number in contacts:
        ws.append([name, number])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_file_bytes(contacts: List[Tuple[str, str]], fmt: str) -> Tuple[bytes, str]:
    if fmt == "vcf":
        return build_vcf_bytes(contacts), "vcf"
    elif fmt == "csv":
        return build_csv_bytes(contacts), "csv"
    elif fmt in ("xlsx", "excel"):
        return build_xlsx_bytes(contacts), "xlsx"
    else:
        return build_txt_bytes(contacts), "txt"


# ─────────────────────────────────────────────────────────────────────────────
# REPLY KEYBOARD BUILDERS
# ─────────────────────────────────────────────────────────────────────────────

def main_menu_keyboard() -> ReplyKeyboardMarkup:
    keyboard = [
        [BTN_ANALYSIS, BTN_CONVERTER],
        [BTN_QVCF, BTN_MAKER],
        [BTN_SPLIT, BTN_MERGE],
        [BTN_EDITOR, BTN_LISTMAKER],
        [BTN_RENFILE, BTN_RENCONT],
        [BTN_SETTINGS, BTN_RESET],
        [BTN_HELP],
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)


def cancel_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [[BTN_CANCEL]],
        resize_keyboard=True,
        one_time_keyboard=True,
    )


def back_cancel_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [[BTN_BACK, BTN_CANCEL]],
        resize_keyboard=True,
        one_time_keyboard=True,
    )


def yes_no_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [[BTN_YES, BTN_NO]],
        resize_keyboard=True,
        one_time_keyboard=True,
    )


def yes_no_reset_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [[BTN_YES_RESET, BTN_NO_CANCEL]],
        resize_keyboard=True,
        one_time_keyboard=True,
    )


def finish_cancel_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [[BTN_FINISH, BTN_CANCEL]],
        resize_keyboard=True,
        one_time_keyboard=True,
    )


def add_more_finish_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [[BTN_ADD_MORE, BTN_GENERATE], [BTN_CANCEL]],
        resize_keyboard=True,
        one_time_keyboard=True,
    )


def upload_more_finish_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [[BTN_FINISH_UPLOAD], [BTN_CANCEL]],
        resize_keyboard=True,
        one_time_keyboard=True,
    )


def merge_finish_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [[BTN_FINISH_MERGING], [BTN_CANCEL]],
        resize_keyboard=True,
        one_time_keyboard=True,
    )


def format_select_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [[BTN_FMT_VCF, BTN_FMT_TXT], [BTN_FMT_EXCEL, BTN_FMT_CSV], [BTN_BACK, BTN_CANCEL]],
        resize_keyboard=True,
        one_time_keyboard=True,
    )


def output_mode_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [[BTN_SINGLE, BTN_SEPARATE], [BTN_CANCEL]],
        resize_keyboard=True,
        one_time_keyboard=True,
    )


def editor_menu_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [[BTN_REMOVE, BTN_EDIT], [BTN_ADD, BTN_DOWNLOAD], [BTN_CANCEL]],
        resize_keyboard=True,
    )


def list_upload_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [[BTN_FINISH_GENERATE], [BTN_CANCEL]],
        resize_keyboard=True,
        one_time_keyboard=True,
    )


def settings_edit_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [
            [BTN_EDIT_1, BTN_EDIT_2, BTN_EDIT_3],
            [BTN_EDIT_4, BTN_EDIT_5, BTN_EDIT_6],
            [BTN_EDIT_7],
            [BTN_BACK, BTN_CANCEL],
        ],
        resize_keyboard=True,
    )


def maker_confirm_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [
            [BTN_EDIT_1, BTN_EDIT_2],
            [BTN_EDIT_3, BTN_EDIT_4],
            [BTN_EDIT_5, BTN_EDIT_6],
            [BTN_EDIT_7],
            [BTN_MAKER_CONFIRM],
            [BTN_CANCEL],
        ],
        resize_keyboard=True,
    )


# ─────────────────────────────────────────────────────────────────────────────
# HELPER: send file safely
# ─────────────────────────────────────────────────────────────────────────────

async def send_file(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    data: bytes,
    filename: str,
    caption: str = "",
) -> None:
    buf = io.BytesIO(data)
    buf.name = filename
    buf.seek(0)
    chat_id = update.effective_chat.id
    await context.bot.send_document(
        chat_id=chat_id,
        document=InputFile(buf, filename=filename),
        caption=caption,
    )


async def go_main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    uid = update.effective_user.id
    reset_udata(uid)
    first = update.effective_user.first_name or "User"
    text = (
        f"📱 *VCF Master Bot*\n\n"
        f"👋 Welcome back, *{first}*!\n"
        f"🆔 Your ID: `{uid}`\n\n"
        f"Choose a feature from the menu below:"
    )
    await update.effective_message.reply_text(
        text,
        reply_markup=main_menu_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return MAIN_MENU


# ─────────────────────────────────────────────────────────────────────────────
# /start
# ─────────────────────────────────────────────────────────────────────────────

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user = update.effective_user
    uid = user.id
    reset_udata(uid)
    first = user.first_name or "User"
    text = (
        f"📱 *VCF Master Bot*\n\n"
        f"👋 Hello, *{first}*!\n"
        f"🆔 Your ID: `{uid}`\n\n"
        f"*Available Features:*\n"
        f"📊 File Analysis — analyze VCF/TXT/CSV files\n"
        f"🔄 File Converter — convert between formats\n"
        f"⚡ Quick VCF — create VCF quickly\n"
        f"🏭 VCF Maker — bulk VCF generation\n"
        f"✂️ Split File — split into smaller files\n"
        f"🔗 Merge Files — merge multiple files\n"
        f"✏️ File Editor — edit contacts in a file\n"
        f"📋 List Maker — extract pending lists from screenshots\n"
        f"📝 Rename File — rename uploaded file\n"
        f"📛 Rename Contact — rename contacts in VCF\n"
        f"⚙️ Settings — configure default values\n"
        f"🔄 Reset — reset settings to defaults\n"
        f"❓ Help — detailed usage instructions\n\n"
        f"👇 *Tap a feature to get started:*"
    )
    await update.message.reply_text(
        text,
        reply_markup=main_menu_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return MAIN_MENU


# ─────────────────────────────────────────────────────────────────────────────
# CANCEL HANDLER
# ─────────────────────────────────────────────────────────────────────────────

async def cancel_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    uid = update.effective_user.id
    reset_udata(uid)
    await update.message.reply_text(
        "❌ Operation cancelled. Returning to main menu.",
        reply_markup=main_menu_keyboard(),
    )
    return MAIN_MENU


# ─────────────────────────────────────────────────────────────────────────────
# MAIN MENU ROUTER
# ─────────────────────────────────────────────────────────────────────────────

async def main_menu_router(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()

    if text == BTN_ANALYSIS:
        return await analysis_start(update, context)
    elif text == BTN_CONVERTER:
        return await converter_start(update, context)
    elif text == BTN_QVCF:
        return await qvcf_start(update, context)
    elif text == BTN_MAKER:
        return await maker_start(update, context)
    elif text == BTN_SPLIT:
        return await split_start(update, context)
    elif text == BTN_MERGE:
        return await merge_start(update, context)
    elif text == BTN_EDITOR:
        return await editor_start(update, context)
    elif text == BTN_LISTMAKER:
        return await listmaker_start(update, context)
    elif text == BTN_RENFILE:
        return await renfile_start(update, context)
    elif text == BTN_RENCONT:
        return await rencont_start(update, context)
    elif text == BTN_SETTINGS:
        return await settings_start(update, context)
    elif text == BTN_RESET:
        return await reset_start(update, context)
    elif text == BTN_HELP:
        return await help_handler(update, context)
    else:
        await update.message.reply_text(
            "⚠️ Please choose a feature from the menu below:",
            reply_markup=main_menu_keyboard(),
        )
        return MAIN_MENU


# ─────────────────────────────────────────────────────────────────────────────
# ① FILE ANALYSIS
# ─────────────────────────────────────────────────────────────────────────────

async def analysis_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    uid = update.effective_user.id
    reset_udata(uid)
    await update.message.reply_text(
        "📊 *File Analysis*\n\n"
        "Please upload your file (VCF, TXT, or CSV).\n"
        "I'll analyze and report:\n"
        "• Total numbers\n"
        "• Numbers by country\n"
        "• Duplicate count\n"
        "• Invalid/junk numbers\n\n"
        "Press ❌ Cancel to go back.",
        reply_markup=back_cancel_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return ANALYSIS_UPLOAD


async def analysis_upload_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text or ""
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)
    if text == BTN_BACK:
        return await go_main_menu(update, context)

    doc = update.message.document
    if not doc:
        await update.message.reply_text(
            "⚠️ Please upload a valid file (VCF/TXT/CSV), or press Cancel.",
            reply_markup=back_cancel_keyboard(),
        )
        return ANALYSIS_UPLOAD

    await update.message.reply_text(
        "⏳ Analyzing your file, please wait...",
        reply_markup=ReplyKeyboardRemove(),
    )

    file = await doc.get_file()
    raw = await file.download_as_bytearray()
    data = bytes(raw)
    filename = doc.file_name or "file.vcf"

    contacts = parse_file_bytes(data, filename)
    total = len(contacts)

    numbers = [c[1] for c in contacts if c[1]]
    seen = set()
    dupes = 0
    country_map: Dict[str, int] = defaultdict(int)
    valid_count = 0
    invalid_count = 0

    for num in numbers:
        is_dup = num in seen
        seen.add(num)
        if is_dup:
            dupes += 1
        if is_valid_number(num):
            valid_count += 1
            cc = get_country_code_str(num) or "XX"
            country_map[cc] += 1
        else:
            invalid_count += 1

    country_lines = []
    for cc, cnt in sorted(country_map.items(), key=lambda x: -x[1]):
        flag = get_flag(cc)
        try:
            pn_obj = phonenumbers.parse("+" + str(phonenumbers.country_code_for_region(cc)))
            country_name = geocoder.country_name_for_number(pn_obj, "en")
        except Exception:
            country_name = cc
        country_lines.append(f"  {flag} {country_name} ({cc}): {cnt}")

    country_section = "\n".join(country_lines) if country_lines else "  🌍 Unknown"

    result = (
        f"📊 *File Analysis Report*\n"
        f"📄 File: `{filename}`\n\n"
        f"📇 *Total Contacts:* {total}\n"
        f"✅ *Valid Numbers:* {valid_count}\n"
        f"❌ *Invalid/Junk:* {invalid_count}\n"
        f"🔁 *Duplicates:* {dupes}\n\n"
        f"🌍 *Numbers by Country:*\n{country_section}\n\n"
        f"_Analysis complete! Choose another feature:_"
    )
    await update.message.reply_text(
        result,
        parse_mode=ParseMode.MARKDOWN,
        reply_markup=main_menu_keyboard(),
    )
    reset_udata(uid)
    return MAIN_MENU


# ─────────────────────────────────────────────────────────────────────────────
# ② FILE CONVERTER
# ─────────────────────────────────────────────────────────────────────────────

async def converter_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    uid = update.effective_user.id
    reset_udata(uid)
    await update.message.reply_text(
        "🔄 *File Converter*\n\n"
        "Please upload your file (VCF, TXT, CSV, or XLSX).\n"
        "I'll ask which format you want to convert it to.\n\n"
        "Press ❌ Cancel to go back.",
        reply_markup=back_cancel_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return CONVERT_UPLOAD


async def converter_upload_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text or ""
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)
    if text == BTN_BACK:
        return await go_main_menu(update, context)

    doc = update.message.document
    if not doc:
        await update.message.reply_text(
            "⚠️ Please upload a valid file, or press Cancel.",
            reply_markup=back_cancel_keyboard(),
        )
        return CONVERT_UPLOAD

    uid = update.effective_user.id
    file = await doc.get_file()
    raw = await file.download_as_bytearray()
    udata = get_udata(uid)
    udata["conv_data"] = bytes(raw)
    udata["conv_filename"] = doc.file_name or "file"

    await update.message.reply_text(
        f"✅ File received: `{doc.file_name}`\n\n"
        "Choose output format:",
        reply_markup=format_select_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return CONVERT_FORMAT


async def converter_format_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    uid = update.effective_user.id

    if text == BTN_CANCEL:
        return await cancel_handler(update, context)
    if text == BTN_BACK:
        return await converter_start(update, context)

    fmt_map = {
        BTN_FMT_TXT: "txt",
        BTN_FMT_VCF: "vcf",
        BTN_FMT_EXCEL: "xlsx",
        BTN_FMT_CSV: "csv",
    }
    fmt = fmt_map.get(text)
    if not fmt:
        await update.message.reply_text(
            "⚠️ Please choose a format from the buttons below:",
            reply_markup=format_select_keyboard(),
        )
        return CONVERT_FORMAT

    udata = get_udata(uid)
    raw = udata.get("conv_data")
    orig_filename = udata.get("conv_filename", "file")
    if not raw:
        await update.message.reply_text(
            "⚠️ File not found. Please restart.",
            reply_markup=main_menu_keyboard(),
        )
        return MAIN_MENU

    contacts = parse_file_bytes(raw, orig_filename)
    if not contacts:
        await update.message.reply_text(
            "⚠️ No contacts found in the file.",
            reply_markup=back_cancel_keyboard(),
        )
        return CONVERT_FORMAT

    file_bytes, ext = build_file_bytes(contacts, fmt)
    base = orig_filename.rsplit(".", 1)[0]
    out_name = f"{base}_converted.{ext}"

    await update.message.reply_text(
        f"⏳ Converting to {fmt.upper()}...",
        reply_markup=ReplyKeyboardRemove(),
    )
    await send_file(
        update, context, file_bytes, out_name,
        caption=f"✅ Converted `{orig_filename}` → `{out_name}`\n📇 {len(contacts)} contacts",
    )
    reset_udata(uid)
    await update.message.reply_text(
        "📱 *VCF Master Bot* — What would you like to do next?",
        reply_markup=main_menu_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return MAIN_MENU


# ─────────────────────────────────────────────────────────────────────────────
# ③ QUICK VCF
# ─────────────────────────────────────────────────────────────────────────────

async def qvcf_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    uid = update.effective_user.id
    reset_udata(uid)
    get_udata(uid)["qvcf_contacts"] = []
    await update.message.reply_text(
        "⚡ *Quick VCF*\n\n"
        "Enter the VCF *file name* (without extension):",
        reply_markup=back_cancel_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return QVCF_FILENAME


async def qvcf_filename(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)
    if text == BTN_BACK:
        return await go_main_menu(update, context)

    uid = update.effective_user.id
    if not text:
        await update.message.reply_text(
            "⚠️ File name cannot be empty. Try again:",
            reply_markup=back_cancel_keyboard(),
        )
        return QVCF_FILENAME
    get_udata(uid)["qvcf_filename"] = text
    await update.message.reply_text(
        f"✅ File name set: `{text}.vcf`\n\nEnter the *contact name*:",
        reply_markup=cancel_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return QVCF_CNAME


async def qvcf_cname(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)

    uid = update.effective_user.id
    if not text:
        await update.message.reply_text(
            "⚠️ Contact name cannot be empty. Try again:",
            reply_markup=cancel_keyboard(),
        )
        return QVCF_CNAME
    get_udata(uid)["qvcf_current_name"] = text
    await update.message.reply_text(
        f"👤 Name: *{text}*\n\nEnter the *phone number* (with country code, e.g. +1234567890):",
        reply_markup=cancel_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return QVCF_CNUMBER


async def qvcf_cnumber(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)

    uid = update.effective_user.id
    if not is_valid_number(text):
        await update.message.reply_text(
            "⚠️ Invalid phone number. Please enter a valid number with country code (e.g. +1234567890):",
            reply_markup=cancel_keyboard(),
        )
        return QVCF_CNUMBER

    udata = get_udata(uid)
    name = udata.get("qvcf_current_name", "Contact")
    number = format_e164(text)
    udata["qvcf_contacts"].append((name, number))
    total = len(udata["qvcf_contacts"])

    await update.message.reply_text(
        f"✅ Added: *{name}* — `{number}`\n📇 Total contacts: {total}\n\nAdd another contact?",
        reply_markup=add_more_finish_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return QVCF_MORE


async def qvcf_more_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)

    if text == BTN_ADD_MORE:
        await update.message.reply_text(
            "Enter the *contact name*:",
            reply_markup=cancel_keyboard(),
            parse_mode=ParseMode.MARKDOWN,
        )
        return QVCF_CNAME

    if text == BTN_GENERATE:
        return await qvcf_generate(update, context)

    await update.message.reply_text(
        "⚠️ Please use the buttons to continue:",
        reply_markup=add_more_finish_keyboard(),
    )
    return QVCF_MORE


async def qvcf_generate(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    uid = update.effective_user.id
    udata = get_udata(uid)
    contacts = udata.get("qvcf_contacts", [])
    fname = udata.get("qvcf_filename", "Contacts")

    if not contacts:
        await update.message.reply_text(
            "⚠️ No contacts added. Returning to menu.",
            reply_markup=main_menu_keyboard(),
        )
        return MAIN_MENU

    vcf_bytes = build_vcf_bytes(contacts)
    out_name = f"{fname}.vcf"
    await update.message.reply_text(
        f"⏳ Generating `{out_name}`...",
        reply_markup=ReplyKeyboardRemove(),
    )
    await send_file(
        update, context, vcf_bytes, out_name,
        caption=f"✅ Quick VCF generated!\n📇 {len(contacts)} contacts in `{out_name}`",
    )
    reset_udata(uid)
    await update.message.reply_text(
        "📱 *VCF Master Bot* — What would you like to do next?",
        reply_markup=main_menu_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return MAIN_MENU


# ─────────────────────────────────────────────────────────────────────────────
# ④ VCF MAKER (FULL WIZARD)
# ─────────────────────────────────────────────────────────────────────────────

SETTINGS_KEYS = [
    ("vcf_filename", "VCF File Name"),
    ("contact_name", "Contact Name"),
    ("per_file", "Contacts per File"),
    ("num_start", "Contact Number Start"),
    ("file_start", "File Number Start"),
    ("country_code", "Country Code"),
    ("group_name", "Group Name"),
]

MAKER_EDIT_BTN_MAP = {
    BTN_EDIT_1: "vcf_filename",
    BTN_EDIT_2: "contact_name",
    BTN_EDIT_3: "per_file",
    BTN_EDIT_4: "num_start",
    BTN_EDIT_5: "file_start",
    BTN_EDIT_6: "country_code",
    BTN_EDIT_7: "group_name",
}

MAKER_EDIT_LABEL_MAP = {
    "vcf_filename": "VCF File Name",
    "contact_name": "Contact Name",
    "per_file": "Contacts per File",
    "num_start": "Contact Number Start",
    "file_start": "File Number Start",
    "country_code": "Country Code",
    "group_name": "Group Name",
}


async def maker_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    uid = update.effective_user.id
    reset_udata(uid)
    settings = get_settings(uid)
    udata = get_udata(uid)
    udata["maker"] = {
        "vcf_filename": settings["vcf_filename"],
        "contact_name": settings["contact_name"],
        "per_file": settings["per_file"],
        "num_start": settings["num_start"],
        "file_start": settings["file_start"],
        "country_code": settings["country_code"],
        "group_name": settings["group_name"],
        "numbers": [],
        "output_mode": "separate",
    }
    await update.message.reply_text(
        "🏭 *VCF Maker*\n\nEnter the *VCF file name* (base name, without extension):",
        reply_markup=back_cancel_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return MAKER_FILENAME


async def maker_filename(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)
    if text == BTN_BACK:
        return await go_main_menu(update, context)

    uid = update.effective_user.id
    if not text:
        await update.message.reply_text(
            "⚠️ File name cannot be empty. Try again:",
            reply_markup=back_cancel_keyboard(),
        )
        return MAKER_FILENAME
    get_udata(uid)["maker"]["vcf_filename"] = text
    await update.message.reply_text(
        f"✅ Base file name: `{text}`\n\n"
        "Enter the *contact name* (base name for all contacts, e.g. \"Customer\"):",
        reply_markup=cancel_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return MAKER_CNAME


async def maker_cname(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)

    uid = update.effective_user.id
    if not text:
        await update.message.reply_text(
            "⚠️ Contact name cannot be empty. Try again:",
            reply_markup=cancel_keyboard(),
        )
        return MAKER_CNAME
    get_udata(uid)["maker"]["contact_name"] = text
    await update.message.reply_text(
        f"✅ Contact name: `{text}`\n\nHow many contacts *per VCF file*? (e.g. 1000):",
        reply_markup=cancel_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return MAKER_PER_FILE


async def maker_per_file(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)

    uid = update.effective_user.id
    if not text.isdigit() or int(text) < 1:
        await update.message.reply_text(
            "⚠️ Please enter a valid positive integer:",
            reply_markup=cancel_keyboard(),
        )
        return MAKER_PER_FILE
    get_udata(uid)["maker"]["per_file"] = int(text)
    await update.message.reply_text(
        f"✅ Contacts per file: `{text}`\n\n"
        "What number should contact numbering *start from*? (e.g. 1 or 501):",
        reply_markup=cancel_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return MAKER_NUM_START


async def maker_num_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)

    uid = update.effective_user.id
    if not text.isdigit():
        await update.message.reply_text(
            "⚠️ Please enter a valid non-negative integer:",
            reply_markup=cancel_keyboard(),
        )
        return MAKER_NUM_START
    get_udata(uid)["maker"]["num_start"] = int(text)
    await update.message.reply_text(
        f"✅ Contact numbering starts from: `{text}`\n\n"
        "What number should *file numbering* start from? (e.g. 1):",
        reply_markup=cancel_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return MAKER_FILE_START


async def maker_file_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)

    uid = update.effective_user.id
    if not text.isdigit() or int(text) < 1:
        await update.message.reply_text(
            "⚠️ Please enter a valid positive integer:",
            reply_markup=cancel_keyboard(),
        )
        return MAKER_FILE_START
    get_udata(uid)["maker"]["file_start"] = int(text)
    await update.message.reply_text(
        "✅ File numbering start set.\n\n"
        "Enter the *country code* for numbers (e.g. `+1` for US, `+44` for UK).\n"
        "Type `auto` for auto-detection:",
        reply_markup=cancel_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return MAKER_COUNTRY


async def maker_country(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)

    uid = update.effective_user.id
    get_udata(uid)["maker"]["country_code"] = text
    await update.message.reply_text(
        f"✅ Country code: `{text}`\n\n"
        "Enter the *group name* (appended after contact name, e.g. \"VIP\"):",
        reply_markup=cancel_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return MAKER_GROUP


async def maker_group(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)

    uid = update.effective_user.id
    get_udata(uid)["maker"]["group_name"] = text
    return await maker_show_confirm(update, context)


async def maker_show_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    uid = update.effective_user.id
    m = get_udata(uid)["maker"]
    summary = (
        f"🏭 *VCF Maker — Confirm Settings*\n\n"
        f"1️⃣ VCF File Name: `{m['vcf_filename']}`\n"
        f"2️⃣ Contact Name: `{m['contact_name']}`\n"
        f"3️⃣ Contacts per File: `{m['per_file']}`\n"
        f"4️⃣ Contact Number Start: `{m['num_start']}`\n"
        f"5️⃣ File Number Start: `{m['file_start']}`\n"
        f"6️⃣ Country Code: `{m['country_code']}`\n"
        f"7️⃣ Group Name: `{m['group_name']}`\n\n"
        f"_Contact format: {m['contact_name']} {m['group_name']} 1_\n"
        f"_File format: {m['vcf_filename']}_1.vcf_\n\n"
        "Tap a numbered button to edit, or *Confirm* to proceed."
    )
    await update.effective_message.reply_text(
        summary,
        reply_markup=maker_confirm_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return MAKER_CONFIRM


async def maker_confirm_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    uid = update.effective_user.id

    if text == BTN_CANCEL:
        return await cancel_handler(update, context)

    if text == BTN_MAKER_CONFIRM:
        get_udata(uid)["maker"]["numbers"] = []
        await update.message.reply_text(
            "📂 *Upload Numbers File(s)*\n\n"
            "Upload your file(s) containing phone numbers (TXT/VCF/CSV/XLSX).\n"
            "You can upload multiple files.\n"
            "When done, tap *✅ Finish Uploading*.",
            reply_markup=upload_more_finish_keyboard(),
            parse_mode=ParseMode.MARKDOWN,
        )
        return MAKER_UPLOAD

    # Check if tapping an edit button
    if text in MAKER_EDIT_BTN_MAP:
        fkey = MAKER_EDIT_BTN_MAP[text]
        flabel = MAKER_EDIT_LABEL_MAP[fkey]
        get_udata(uid)["maker_edit_field"] = fkey
        current_val = get_udata(uid)["maker"].get(fkey, "")
        await update.message.reply_text(
            f"✏️ Enter new value for *{flabel}*:\n_(Current: `{current_val}`)_",
            reply_markup=back_cancel_keyboard(),
            parse_mode=ParseMode.MARKDOWN,
        )
        return MAKER_EDIT_FIELD

    await update.message.reply_text(
        "⚠️ Please use the buttons to proceed:",
        reply_markup=maker_confirm_keyboard(),
    )
    return MAKER_CONFIRM


async def maker_edit_field_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    uid = update.effective_user.id

    if text == BTN_CANCEL:
        return await cancel_handler(update, context)
    if text == BTN_BACK:
        return await maker_show_confirm(update, context)

    udata = get_udata(uid)
    fkey = udata.get("maker_edit_field")
    if fkey in ("per_file", "num_start", "file_start"):
        if not text.isdigit():
            await update.message.reply_text(
                "⚠️ Please enter a valid integer:",
                reply_markup=back_cancel_keyboard(),
            )
            return MAKER_EDIT_FIELD
        udata["maker"][fkey] = int(text)
    else:
        udata["maker"][fkey] = text

    return await maker_show_confirm(update, context)


async def maker_upload_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text or ""
    uid = update.effective_user.id

    if text == BTN_CANCEL:
        return await cancel_handler(update, context)

    if text == BTN_FINISH_UPLOAD:
        udata = get_udata(uid)
        numbers = udata["maker"].get("numbers", [])
        if not numbers:
            await update.message.reply_text(
                "⚠️ No numbers uploaded yet. Please upload at least one file.",
                reply_markup=upload_more_finish_keyboard(),
            )
            return MAKER_UPLOAD

        await update.message.reply_text(
            f"📊 Total numbers loaded: *{len(numbers)}*\n\n"
            f"How would you like the output?\n"
            f"• *Single Merged File* — all contacts in one VCF\n"
            f"• *Separate Files* — split by your per-file setting ({udata['maker']['per_file']} per file)",
            reply_markup=output_mode_keyboard(),
            parse_mode=ParseMode.MARKDOWN,
        )
        return MAKER_OUTPUT_MODE

    doc = update.message.document
    if not doc:
        await update.message.reply_text(
            "⚠️ Please upload a file, or tap *✅ Finish Uploading*.",
            reply_markup=upload_more_finish_keyboard(),
            parse_mode=ParseMode.MARKDOWN,
        )
        return MAKER_UPLOAD

    file = await doc.get_file()
    raw = await file.download_as_bytearray()
    data = bytes(raw)
    filename = doc.file_name or "numbers.txt"
    contacts = parse_file_bytes(data, filename)
    numbers_extracted = [c[1] for c in contacts if c[1]]
    get_udata(uid)["maker"]["numbers"].extend(numbers_extracted)
    total = len(get_udata(uid)["maker"]["numbers"])

    await update.message.reply_text(
        f"✅ Loaded `{filename}` — {len(numbers_extracted)} numbers.\n"
        f"📊 Total so far: *{total}* numbers.\n\n"
        "Upload another file or tap *✅ Finish Uploading*.",
        reply_markup=upload_more_finish_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return MAKER_UPLOAD


async def maker_output_mode_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    uid = update.effective_user.id

    if text == BTN_CANCEL:
        return await cancel_handler(update, context)

    if text == BTN_SINGLE:
        get_udata(uid)["maker"]["output_mode"] = "single"
    elif text == BTN_SEPARATE:
        get_udata(uid)["maker"]["output_mode"] = "separate"
    else:
        await update.message.reply_text(
            "⚠️ Please choose an output mode:",
            reply_markup=output_mode_keyboard(),
        )
        return MAKER_OUTPUT_MODE

    await update.message.reply_text(
        "⏳ Generating VCF file(s), please wait...",
        reply_markup=ReplyKeyboardRemove(),
    )
    await maker_generate(update, context)
    return MAIN_MENU


async def maker_generate(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    uid = update.effective_user.id
    udata = get_udata(uid)
    m = udata["maker"]
    numbers = m["numbers"]
    contact_name = m["contact_name"]
    group_name = m["group_name"]
    per_file = m["per_file"]
    num_start = m["num_start"]
    file_start = m["file_start"]
    vcf_filename = m["vcf_filename"]
    country_code = m["country_code"]
    output_mode = m["output_mode"]

    all_contacts: List[Tuple[str, str]] = []
    for i, raw_num in enumerate(numbers):
        idx = num_start + i
        name = f"{contact_name} {group_name} {idx}" if group_name else f"{contact_name} {idx}"
        num = raw_num
        if country_code.lower() != "auto" and not num.startswith("+"):
            cc_clean = country_code.lstrip("+")
            num = cc_clean + num.lstrip("0")
        if not num.startswith("+"):
            num = "+" + num
        all_contacts.append((name, num))

    chat_id = update.effective_chat.id

    if output_mode == "single":
        vcf_bytes = build_vcf_bytes(all_contacts)
        out_name = f"{vcf_filename}.vcf"
        await send_file(
            update, context, vcf_bytes, out_name,
            caption=f"✅ VCF Maker complete!\n📇 {len(all_contacts)} contacts in `{out_name}`",
        )
    else:
        chunks = [all_contacts[i:i + per_file] for i in range(0, len(all_contacts), per_file)]
        sent = 0
        for ci, chunk in enumerate(chunks):
            fnum = file_start + ci
            out_name = f"{vcf_filename}_{fnum}.vcf"
            vcf_bytes = build_vcf_bytes(chunk)
            caption = f"📂 File {ci + 1}/{len(chunks)}: `{out_name}` — {len(chunk)} contacts"
            await send_file(update, context, vcf_bytes, out_name, caption=caption)
            sent += 1

        await context.bot.send_message(
            chat_id=chat_id,
            text=f"✅ VCF Maker complete!\n📂 {sent} file(s) generated\n📇 {len(all_contacts)} total contacts",
        )

    reset_udata(uid)
    await context.bot.send_message(
        chat_id=chat_id,
        text="📱 *VCF Master Bot* — What would you like to do next?",
        reply_markup=main_menu_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )


# ─────────────────────────────────────────────────────────────────────────────
# ⑤ SPLIT FILE
# ─────────────────────────────────────────────────────────────────────────────

async def split_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    uid = update.effective_user.id
    reset_udata(uid)
    await update.message.reply_text(
        "✂️ *Split File*\n\nUpload your file (VCF, TXT, or CSV) to split.\n\n"
        "Press ❌ Cancel to go back.",
        reply_markup=back_cancel_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return SPLIT_UPLOAD


async def split_upload_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text or ""
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)
    if text == BTN_BACK:
        return await go_main_menu(update, context)

    uid = update.effective_user.id
    doc = update.message.document
    if not doc:
        await update.message.reply_text(
            "⚠️ Please upload a valid file, or press Cancel.",
            reply_markup=back_cancel_keyboard(),
        )
        return SPLIT_UPLOAD

    file = await doc.get_file()
    raw = await file.download_as_bytearray()
    udata = get_udata(uid)
    udata["split_data"] = bytes(raw)
    udata["split_filename"] = doc.file_name or "file.vcf"

    await update.message.reply_text(
        f"✅ File received: `{doc.file_name}`\n\nHow many contacts *per split file*? (e.g. 50):",
        reply_markup=cancel_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return SPLIT_COUNT


async def split_count_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)

    uid = update.effective_user.id
    if not text.isdigit() or int(text) < 1:
        await update.message.reply_text(
            "⚠️ Please enter a valid positive integer:",
            reply_markup=cancel_keyboard(),
        )
        return SPLIT_COUNT

    udata = get_udata(uid)
    count = int(text)
    raw = udata.get("split_data")
    filename = udata.get("split_filename", "file.vcf")
    fmt = detect_format(filename)
    contacts = parse_file_bytes(raw, filename)

    if not contacts:
        await update.message.reply_text(
            "⚠️ No contacts found in the file.",
            reply_markup=back_cancel_keyboard(),
        )
        return SPLIT_UPLOAD

    chunks = [contacts[i:i + count] for i in range(0, len(contacts), count)]
    base = filename.rsplit(".", 1)[0]
    ext = fmt if fmt else "txt"

    await update.message.reply_text(
        f"⏳ Splitting into {len(chunks)} files ({count} contacts each)...",
        reply_markup=ReplyKeyboardRemove(),
    )
    for i, chunk in enumerate(chunks):
        out_bytes, out_ext = build_file_bytes(chunk, ext)
        out_name = f"{base}_part{i + 1}.{out_ext}"
        await send_file(
            update, context, out_bytes, out_name,
            caption=f"✂️ Part {i + 1}/{len(chunks)}: `{out_name}` — {len(chunk)} contacts",
        )

    await update.message.reply_text(
        f"✅ Split complete!\n📂 {len(chunks)} files created from {len(contacts)} contacts.",
        reply_markup=main_menu_keyboard(),
    )
    reset_udata(uid)
    return MAIN_MENU


# ─────────────────────────────────────────────────────────────────────────────
# ⑥ MERGE FILES
# ─────────────────────────────────────────────────────────────────────────────

async def merge_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    uid = update.effective_user.id
    reset_udata(uid)
    udata = get_udata(uid)
    udata["merge_contacts"] = []
    udata["merge_fmt"] = "vcf"

    await update.message.reply_text(
        "🔗 *Merge Files*\n\nUpload your files one by one.\n"
        "When done, tap *✅ Finish Merging*.",
        reply_markup=merge_finish_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return MERGE_UPLOAD


async def merge_upload_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text or ""
    uid = update.effective_user.id

    if text == BTN_CANCEL:
        return await cancel_handler(update, context)

    if text == BTN_FINISH_MERGING:
        udata = get_udata(uid)
        contacts = udata.get("merge_contacts", [])
        fmt = udata.get("merge_fmt", "vcf")

        if not contacts:
            await update.message.reply_text(
                "⚠️ No files uploaded yet.",
                reply_markup=merge_finish_keyboard(),
            )
            return MERGE_UPLOAD

        out_bytes, ext = build_file_bytes(contacts, fmt)
        out_name = f"merged.{ext}"
        await update.message.reply_text(
            f"⏳ Merging {len(contacts)} contacts...",
            reply_markup=ReplyKeyboardRemove(),
        )
        await send_file(
            update, context, out_bytes, out_name,
            caption=f"✅ Merge complete!\n📇 {len(contacts)} contacts in `{out_name}`",
        )
        reset_udata(uid)
        await update.message.reply_text(
            "📱 *VCF Master Bot* — What would you like to do next?",
            reply_markup=main_menu_keyboard(),
            parse_mode=ParseMode.MARKDOWN,
        )
        return MAIN_MENU

    doc = update.message.document
    if not doc:
        await update.message.reply_text(
            "⚠️ Please upload a file, or tap *✅ Finish Merging*.",
            reply_markup=merge_finish_keyboard(),
            parse_mode=ParseMode.MARKDOWN,
        )
        return MERGE_UPLOAD

    file = await doc.get_file()
    raw = await file.download_as_bytearray()
    data = bytes(raw)
    filename = doc.file_name or "file.vcf"
    contacts = parse_file_bytes(data, filename)
    udata = get_udata(uid)
    if not udata["merge_contacts"]:
        udata["merge_fmt"] = detect_format(filename) or "vcf"
    udata["merge_contacts"].extend(contacts)
    total = len(udata["merge_contacts"])

    await update.message.reply_text(
        f"✅ Loaded `{filename}` — {len(contacts)} contacts.\n"
        f"📊 Total so far: *{total}* contacts.\n\n"
        "Upload another file or tap *✅ Finish Merging*.",
        reply_markup=merge_finish_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return MERGE_UPLOAD


# ─────────────────────────────────────────────────────────────────────────────
# ⑦ FILE EDITOR
# ─────────────────────────────────────────────────────────────────────────────

async def editor_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    uid = update.effective_user.id
    reset_udata(uid)
    await update.message.reply_text(
        "✏️ *File Editor*\n\nUpload a file (VCF, TXT, or CSV) to edit.\n\n"
        "Press ❌ Cancel to go back.",
        reply_markup=back_cancel_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return EDITOR_UPLOAD


async def editor_upload_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text or ""
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)
    if text == BTN_BACK:
        return await go_main_menu(update, context)

    uid = update.effective_user.id
    doc = update.message.document
    if not doc:
        await update.message.reply_text(
            "⚠️ Please upload a valid file, or press Cancel.",
            reply_markup=back_cancel_keyboard(),
        )
        return EDITOR_UPLOAD

    file = await doc.get_file()
    raw = await file.download_as_bytearray()
    data = bytes(raw)
    filename = doc.file_name or "file.vcf"
    contacts = parse_file_bytes(data, filename)
    udata = get_udata(uid)
    udata["editor_contacts"] = list(contacts)
    udata["editor_fmt"] = detect_format(filename) or "vcf"
    udata["editor_filename"] = filename
    return await editor_show_menu(update, context)


async def editor_show_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    uid = update.effective_user.id
    udata = get_udata(uid)
    contacts = udata.get("editor_contacts", [])
    lines = []
    for i, (name, num) in enumerate(contacts[:20]):
        display = f"{i + 1}. {name or 'No Name'} — {num}"
        lines.append(display)
    if len(contacts) > 20:
        lines.append(f"... and {len(contacts) - 20} more")

    contact_list = "\n".join(lines) if lines else "_(empty)_"
    text = (
        f"✏️ *File Editor* — {len(contacts)} contacts\n\n"
        f"```\n{contact_list}\n```\n\n"
        "Choose an action:"
    )
    await update.effective_message.reply_text(
        text,
        reply_markup=editor_menu_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return EDITOR_MENU


async def editor_menu_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    uid = update.effective_user.id

    if text == BTN_CANCEL:
        return await cancel_handler(update, context)

    if text == BTN_REMOVE:
        udata = get_udata(uid)
        contacts = udata.get("editor_contacts", [])
        lines = [
            f"{i + 1}. {name or 'No Name'} — {num}"
            for i, (name, num) in enumerate(contacts[:30])
        ]
        contact_list = "\n".join(lines) if lines else "_(empty)_"
        await update.message.reply_text(
            f"➖ *Remove Contact*\n\nType the *number* (1, 2, 3...) of the contact to remove:\n\n"
            f"```\n{contact_list}\n```",
            reply_markup=back_cancel_keyboard(),
            parse_mode=ParseMode.MARKDOWN,
        )
        return EDITOR_REMOVE

    elif text == BTN_EDIT:
        udata = get_udata(uid)
        contacts = udata.get("editor_contacts", [])
        lines = [
            f"{i + 1}. {name or 'No Name'} — {num}"
            for i, (name, num) in enumerate(contacts[:30])
        ]
        contact_list = "\n".join(lines) if lines else "_(empty)_"
        await update.message.reply_text(
            f"✏️ *Edit Contact*\n\nType the *number* of the contact to edit:\n\n"
            f"```\n{contact_list}\n```",
            reply_markup=back_cancel_keyboard(),
            parse_mode=ParseMode.MARKDOWN,
        )
        return EDITOR_EDIT_SEL

    elif text == BTN_ADD:
        await update.message.reply_text(
            "➕ *Add Contact*\n\nEnter the *name* for the new contact:",
            reply_markup=back_cancel_keyboard(),
            parse_mode=ParseMode.MARKDOWN,
        )
        return EDITOR_ADD_NAME

    elif text == BTN_DOWNLOAD:
        return await editor_finish(update, context)

    await update.message.reply_text(
        "⚠️ Please use the buttons to choose an action:",
        reply_markup=editor_menu_keyboard(),
    )
    return EDITOR_MENU


async def editor_remove_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)
    if text == BTN_BACK:
        return await editor_show_menu(update, context)

    uid = update.effective_user.id
    udata = get_udata(uid)
    contacts = udata.get("editor_contacts", [])
    if not text.isdigit() or int(text) < 1 or int(text) > len(contacts):
        await update.message.reply_text(
            f"⚠️ Enter a number between 1 and {len(contacts)}:",
            reply_markup=back_cancel_keyboard(),
        )
        return EDITOR_REMOVE
    idx = int(text) - 1
    removed = contacts.pop(idx)
    udata["editor_contacts"] = contacts
    await update.message.reply_text(
        f"✅ Removed: *{removed[0] or 'No Name'}* — `{removed[1]}`",
        parse_mode=ParseMode.MARKDOWN,
    )
    return await editor_show_menu(update, context)


async def editor_edit_select_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)
    if text == BTN_BACK:
        return await editor_show_menu(update, context)

    uid = update.effective_user.id
    udata = get_udata(uid)
    contacts = udata.get("editor_contacts", [])
    if not text.isdigit() or int(text) < 1 or int(text) > len(contacts):
        await update.message.reply_text(
            f"⚠️ Enter a number between 1 and {len(contacts)}:",
            reply_markup=back_cancel_keyboard(),
        )
        return EDITOR_EDIT_SEL
    idx = int(text) - 1
    udata["editor_edit_idx"] = idx
    name, num = contacts[idx]
    await update.message.reply_text(
        f"✏️ Editing: *{name or 'No Name'}* — `{num}`\n\n"
        "Enter new *name* (or send `-` to keep current):",
        reply_markup=cancel_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return EDITOR_EDIT_NAME


async def editor_edit_name_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)

    uid = update.effective_user.id
    udata = get_udata(uid)
    idx = udata.get("editor_edit_idx", 0)
    contacts = udata.get("editor_contacts", [])
    old_name, old_num = contacts[idx]
    if text != "-":
        contacts[idx] = (text, old_num)
    udata["editor_contacts"] = contacts
    await update.message.reply_text(
        f"Now enter new *phone number* (or send `-` to keep `{old_num}`):",
        reply_markup=cancel_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return EDITOR_EDIT_NUM


async def editor_edit_num_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)

    uid = update.effective_user.id
    udata = get_udata(uid)
    idx = udata.get("editor_edit_idx", 0)
    contacts = udata.get("editor_contacts", [])
    old_name, old_num = contacts[idx]
    if text != "-":
        if not is_valid_number(text):
            await update.message.reply_text(
                "⚠️ Invalid phone number. Enter again (or `-` to keep):",
                reply_markup=cancel_keyboard(),
            )
            return EDITOR_EDIT_NUM
        contacts[idx] = (old_name, format_e164(text))
    udata["editor_contacts"] = contacts
    new_name, new_num = contacts[idx]
    await update.message.reply_text(
        f"✅ Updated: *{new_name}* — `{new_num}`",
        parse_mode=ParseMode.MARKDOWN,
    )
    return await editor_show_menu(update, context)


async def editor_add_name_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)
    if text == BTN_BACK:
        return await editor_show_menu(update, context)

    uid = update.effective_user.id
    if not text:
        await update.message.reply_text(
            "⚠️ Name cannot be empty:",
            reply_markup=back_cancel_keyboard(),
        )
        return EDITOR_ADD_NAME
    get_udata(uid)["editor_new_name"] = text
    await update.message.reply_text(
        f"👤 Name: *{text}*\n\nEnter phone number:",
        reply_markup=cancel_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return EDITOR_ADD_NUM


async def editor_add_num_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)

    uid = update.effective_user.id
    if not is_valid_number(text):
        await update.message.reply_text(
            "⚠️ Invalid phone number. Try again:",
            reply_markup=cancel_keyboard(),
        )
        return EDITOR_ADD_NUM
    udata = get_udata(uid)
    name = udata.get("editor_new_name", "")
    num = format_e164(text)
    udata["editor_contacts"].append((name, num))
    await update.message.reply_text(
        f"✅ Added: *{name}* — `{num}`",
        parse_mode=ParseMode.MARKDOWN,
    )
    return await editor_show_menu(update, context)


async def editor_finish(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    uid = update.effective_user.id
    udata = get_udata(uid)
    contacts = udata.get("editor_contacts", [])
    fmt = udata.get("editor_fmt", "vcf")
    filename = udata.get("editor_filename", "edited.vcf")
    base = filename.rsplit(".", 1)[0]

    out_bytes, ext = build_file_bytes(contacts, fmt)
    out_name = f"{base}_edited.{ext}"
    await update.effective_message.reply_text(
        f"⏳ Saving {len(contacts)} contacts...",
        reply_markup=ReplyKeyboardRemove(),
    )
    await send_file(
        update, context, out_bytes, out_name,
        caption=f"✅ Edited file ready!\n📇 {len(contacts)} contacts in `{out_name}`",
    )
    reset_udata(uid)
    await update.effective_message.reply_text(
        "📱 *VCF Master Bot* — What would you like to do next?",
        reply_markup=main_menu_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return MAIN_MENU


# ─────────────────────────────────────────────────────────────────────────────
# ⑧ LIST MAKER
# ─────────────────────────────────────────────────────────────────────────────

def _extract_list_entry_from_caption(caption: str) -> Dict[str, Any]:
    if "|" in caption:
        parts = caption.split("|", 1)
        group = parts[0].strip()
        try:
            count = int(re.search(r"\d+", parts[1]).group())
        except Exception:
            count = 0
        return {"group": group, "count": count}
    elif ":" in caption:
        parts = caption.split(":", 1)
        group = parts[0].strip()
        try:
            count = int(re.search(r"\d+", parts[1]).group())
        except Exception:
            count = 0
        return {"group": group, "count": count}
    else:
        m = re.search(r"\d+", caption)
        count = int(m.group()) if m else 0
        group = re.sub(r"\d+", "", caption).strip() or "Unknown Group"
        return {"group": group, "count": count}


async def listmaker_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    uid = update.effective_user.id
    reset_udata(uid)
    await update.message.reply_text(
        "📋 *List Maker*\n\nEnter the *list name* (for the output document):",
        reply_markup=back_cancel_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return LIST_NAME


async def listmaker_name_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)
    if text == BTN_BACK:
        return await go_main_menu(update, context)

    uid = update.effective_user.id
    if not text:
        await update.message.reply_text(
            "⚠️ List name cannot be empty:",
            reply_markup=back_cancel_keyboard(),
        )
        return LIST_NAME
    udata = get_udata(uid)
    udata["list_name"] = text
    udata["list_entries"] = []
    await update.message.reply_text(
        f"✅ List name: *{text}*\n\n"
        "📸 Upload WhatsApp pending request screenshots.\n"
        "I'll extract group names and pending counts.\n"
        "Add a caption in format `GroupName | Count` for auto-extraction.\n"
        "When done, tap *🏁 Generate List*.",
        reply_markup=list_upload_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return LIST_UPLOAD


async def listmaker_upload_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text or ""
    uid = update.effective_user.id

    if text == BTN_CANCEL:
        return await cancel_handler(update, context)

    if text == BTN_FINISH_GENERATE:
        udata = get_udata(uid)
        entries = udata.get("list_entries", [])
        list_name = udata.get("list_name", "List")

        if not entries:
            await update.message.reply_text(
                "⚠️ No screenshots received.\n"
                "📸 Please send photos with captions in format `GroupName | Count`.",
                reply_markup=list_upload_keyboard(),
                parse_mode=ParseMode.MARKDOWN,
            )
            return LIST_UPLOAD

        lines = [f"📋 *{list_name}*\n"]
        total_pending = 0
        for i, entry in enumerate(entries, 1):
            group = entry.get("group", "Unknown")
            count = entry.get("count", 0)
            total_pending += count
            lines.append(f"{i}. {group} — {count} pending")
        lines.append(f"\n📊 *Total Groups:* {len(entries)}")
        lines.append(f"📊 *Total Pending:* {total_pending}")
        result_text = "\n".join(lines)

        await update.message.reply_text(
            result_text,
            reply_markup=ReplyKeyboardRemove(),
            parse_mode=ParseMode.MARKDOWN,
        )
        txt_bytes = result_text.encode("utf-8")
        out_name = f"{list_name.replace(' ', '_')}_list.txt"
        await send_file(
            update, context, txt_bytes, out_name,
            caption=f"📋 {list_name} — {len(entries)} entries",
        )
        reset_udata(uid)
        await update.message.reply_text(
            "📱 *VCF Master Bot* — What would you like to do next?",
            reply_markup=main_menu_keyboard(),
            parse_mode=ParseMode.MARKDOWN,
        )
        return MAIN_MENU

    photo = update.message.photo
    doc = update.message.document

    if photo:
        caption = update.message.caption or ""
        entry = _extract_list_entry_from_caption(caption)
        get_udata(uid)["list_entries"].append(entry)
        total = len(get_udata(uid)["list_entries"])
        await update.message.reply_text(
            f"📸 Screenshot received ({total} so far).\n"
            f"ℹ️ *Tip:* Include caption as `GroupName | Count` for auto-extraction.\n"
            "Upload more or tap *🏁 Generate List*.",
            reply_markup=list_upload_keyboard(),
            parse_mode=ParseMode.MARKDOWN,
        )
        return LIST_UPLOAD

    if doc:
        await update.message.reply_text(
            "⚠️ Please send a photo/screenshot, not a document.",
            reply_markup=list_upload_keyboard(),
        )
        return LIST_UPLOAD

    await update.message.reply_text(
        "⚠️ Please send a screenshot image, or tap *🏁 Generate List*.",
        reply_markup=list_upload_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return LIST_UPLOAD


# ─────────────────────────────────────────────────────────────────────────────
# ⑨ RENAME FILE
# ─────────────────────────────────────────────────────────────────────────────

async def renfile_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    uid = update.effective_user.id
    reset_udata(uid)
    await update.message.reply_text(
        "📝 *Rename File*\n\nUpload the file you want to rename.\n\n"
        "Press ❌ Cancel to go back.",
        reply_markup=back_cancel_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return RENFILE_UPLOAD


async def renfile_upload_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text or ""
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)
    if text == BTN_BACK:
        return await go_main_menu(update, context)

    uid = update.effective_user.id
    doc = update.message.document
    if not doc:
        await update.message.reply_text(
            "⚠️ Please upload a file, or press Cancel.",
            reply_markup=back_cancel_keyboard(),
        )
        return RENFILE_UPLOAD

    file = await doc.get_file()
    raw = await file.download_as_bytearray()
    udata = get_udata(uid)
    udata["renfile_data"] = bytes(raw)
    udata["renfile_orig"] = doc.file_name or "file"
    ext = (doc.file_name or "file").rsplit(".", 1)[-1] if "." in (doc.file_name or "") else ""
    udata["renfile_ext"] = ext

    await update.message.reply_text(
        f"✅ File received: `{doc.file_name}`\n\nEnter the *new file name* (without extension):",
        reply_markup=cancel_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return RENFILE_NAME


async def renfile_name_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)

    uid = update.effective_user.id
    if not text:
        await update.message.reply_text(
            "⚠️ File name cannot be empty:",
            reply_markup=cancel_keyboard(),
        )
        return RENFILE_NAME
    udata = get_udata(uid)
    data = udata.get("renfile_data")
    ext = udata.get("renfile_ext", "")
    out_name = f"{text}.{ext}" if ext else text

    await update.message.reply_text(
        f"⏳ Renaming to `{out_name}`...",
        reply_markup=ReplyKeyboardRemove(),
    )
    await send_file(update, context, data, out_name, caption=f"✅ File renamed to `{out_name}`")
    reset_udata(uid)
    await update.message.reply_text(
        "📱 *VCF Master Bot* — What would you like to do next?",
        reply_markup=main_menu_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return MAIN_MENU


# ─────────────────────────────────────────────────────────────────────────────
# ⑩ RENAME CONTACT
# ─────────────────────────────────────────────────────────────────────────────

async def rencont_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    uid = update.effective_user.id
    reset_udata(uid)
    await update.message.reply_text(
        "📛 *Rename Contact*\n\nUpload a VCF file to rename all contacts.\n\n"
        "Press ❌ Cancel to go back.",
        reply_markup=back_cancel_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return RENCONT_UPLOAD


async def rencont_upload_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text or ""
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)
    if text == BTN_BACK:
        return await go_main_menu(update, context)

    uid = update.effective_user.id
    doc = update.message.document
    if not doc:
        await update.message.reply_text(
            "⚠️ Please upload a VCF file, or press Cancel.",
            reply_markup=back_cancel_keyboard(),
        )
        return RENCONT_UPLOAD

    file = await doc.get_file()
    raw = await file.download_as_bytearray()
    data = bytes(raw)
    filename = doc.file_name or "contacts.vcf"
    contacts = parse_vcf_bytes(data)

    if not contacts:
        await update.message.reply_text(
            "⚠️ No contacts found in VCF file.",
            reply_markup=back_cancel_keyboard(),
        )
        return RENCONT_UPLOAD

    udata = get_udata(uid)
    udata["rencont_contacts"] = contacts
    udata["rencont_filename"] = filename
    await update.message.reply_text(
        f"✅ Loaded `{filename}` — {len(contacts)} contacts.\n\n"
        "Enter the *new base contact name* (numbering will be preserved):\n"
        "_Example: enter \"Client\" → contacts become Client 1, Client 2..._",
        reply_markup=cancel_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return RENCONT_NAME


async def rencont_name_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)

    uid = update.effective_user.id
    if not text:
        await update.message.reply_text(
            "⚠️ Name cannot be empty:",
            reply_markup=cancel_keyboard(),
        )
        return RENCONT_NAME

    udata = get_udata(uid)
    contacts = udata.get("rencont_contacts", [])
    filename = udata.get("rencont_filename", "contacts.vcf")

    renamed = []
    for i, (orig_name, num) in enumerate(contacts, 1):
        m = re.search(r"\d+$", orig_name.strip())
        suffix = m.group() if m else str(i)
        renamed.append((f"{text} {suffix}", num))

    vcf_bytes = build_vcf_bytes(renamed)
    base = filename.rsplit(".", 1)[0]
    out_name = f"{base}_renamed.vcf"

    await update.message.reply_text(
        f"⏳ Renaming {len(renamed)} contacts...",
        reply_markup=ReplyKeyboardRemove(),
    )
    await send_file(
        update, context, vcf_bytes, out_name,
        caption=f"✅ Renamed {len(renamed)} contacts → `{text} N`\n📄 File: `{out_name}`",
    )
    reset_udata(uid)
    await update.message.reply_text(
        "📱 *VCF Master Bot* — What would you like to do next?",
        reply_markup=main_menu_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return MAIN_MENU


# ─────────────────────────────────────────────────────────────────────────────
# ⑪ SETTINGS
# ─────────────────────────────────────────────────────────────────────────────

SETTINGS_EDIT_BTN_MAP = {
    BTN_EDIT_1: "vcf_filename",
    BTN_EDIT_2: "contact_name",
    BTN_EDIT_3: "per_file",
    BTN_EDIT_4: "num_start",
    BTN_EDIT_5: "file_start",
    BTN_EDIT_6: "country_code",
    BTN_EDIT_7: "group_name",
}

SETTINGS_LABEL_MAP = {
    "vcf_filename": "VCF File Name",
    "contact_name": "Contact Name",
    "per_file": "Contacts per File",
    "num_start": "Number Start",
    "file_start": "File Start",
    "country_code": "Country Code",
    "group_name": "Group Name",
}


async def settings_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    return await settings_show(update, context)


async def settings_show(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    uid = update.effective_user.id
    s = get_settings(uid)
    lines = ["⚙️ *Settings*\n"]
    for key, label in SETTINGS_KEYS:
        lines.append(f"• *{label}:* `{s[key]}`")
    lines.append("\nTap a button to edit a setting, or press 🔙 Back.")
    text = "\n".join(lines)
    await update.effective_message.reply_text(
        text,
        reply_markup=settings_edit_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return SETTINGS_MENU


async def settings_menu_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    uid = update.effective_user.id

    if text == BTN_CANCEL:
        return await cancel_handler(update, context)
    if text == BTN_BACK:
        return await go_main_menu(update, context)

    if text in SETTINGS_EDIT_BTN_MAP:
        key = SETTINGS_EDIT_BTN_MAP[text]
        label = SETTINGS_LABEL_MAP.get(key, key)
        get_udata(uid)["settings_edit_key"] = key
        current = get_settings(uid)[key]
        await update.message.reply_text(
            f"⚙️ *Edit Setting: {label}*\n\nCurrent value: `{current}`\n\nEnter new value:",
            reply_markup=back_cancel_keyboard(),
            parse_mode=ParseMode.MARKDOWN,
        )
        return SETTINGS_EDIT

    await update.message.reply_text(
        "⚠️ Please choose a setting to edit:",
        reply_markup=settings_edit_keyboard(),
    )
    return SETTINGS_MENU


async def settings_edit_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_CANCEL:
        return await cancel_handler(update, context)
    if text == BTN_BACK:
        return await settings_show(update, context)

    uid = update.effective_user.id
    udata = get_udata(uid)
    key = udata.get("settings_edit_key")
    if not key:
        return await settings_show(update, context)

    s = get_settings(uid)
    if key in ("per_file", "num_start", "file_start"):
        if not text.isdigit():
            await update.message.reply_text(
                "⚠️ Please enter a valid integer:",
                reply_markup=back_cancel_keyboard(),
            )
            return SETTINGS_EDIT
        s[key] = int(text)
    else:
        s[key] = text

    await update.message.reply_text(
        f"✅ Setting updated: *{SETTINGS_LABEL_MAP.get(key, key)}* = `{text}`",
        parse_mode=ParseMode.MARKDOWN,
    )
    return await settings_show(update, context)


# ─────────────────────────────────────────────────────────────────────────────
# ⑫ RESET
# ─────────────────────────────────────────────────────────────────────────────

async def reset_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text(
        "🔄 *Reset Settings*\n\n"
        "Are you sure you want to reset ALL settings to default values?\n"
        "This cannot be undone.",
        reply_markup=yes_no_reset_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return RESET_CONFIRM


async def reset_confirm_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    uid = update.effective_user.id

    if text == BTN_YES_RESET:
        USER_SETTINGS[uid] = copy.deepcopy(DEFAULT_SETTINGS)
        lines = ["✅ *Settings reset to defaults!*\n"]
        for k, l in SETTINGS_KEYS:
            lines.append(f"• *{l}:* `{DEFAULT_SETTINGS[k]}`")
        await update.message.reply_text(
            "\n".join(lines),
            reply_markup=main_menu_keyboard(),
            parse_mode=ParseMode.MARKDOWN,
        )
        return MAIN_MENU

    if text == BTN_NO_CANCEL:
        await update.message.reply_text(
            "✅ Reset cancelled. Settings unchanged.",
            reply_markup=main_menu_keyboard(),
        )
        return MAIN_MENU

    await update.message.reply_text(
        "⚠️ Please tap Yes or No:",
        reply_markup=yes_no_reset_keyboard(),
    )
    return RESET_CONFIRM


# ─────────────────────────────────────────────────────────────────────────────
# ⑬ HELP
# ─────────────────────────────────────────────────────────────────────────────

async def help_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = (
        "❓ *VCF Master Bot — Help Guide*\n\n"

        "📊 *File Analysis*\n"
        "Upload any VCF/TXT/CSV file. The bot analyzes:\n"
        "total contacts, valid/invalid numbers, duplicates, country breakdown.\n\n"

        "🔄 *File Converter*\n"
        "Upload a file and choose the output format:\n"
        "VCF, TXT, CSV, or Excel (XLSX).\n\n"

        "⚡ *Quick VCF*\n"
        "Quickly create a VCF by entering contacts one by one.\n"
        "Enter file name → contact name → number → add more or finish.\n\n"

        "🏭 *VCF Maker*\n"
        "Bulk VCF generator. Configure: file name, contact name,\n"
        "contacts per file, numbering start, file numbering start,\n"
        "country code, group name. Upload numbers file(s) → get VCF files.\n"
        "Contact format: `ContactName GroupName 1`\n\n"

        "✂️ *Split File*\n"
        "Upload a file and specify how many contacts per part.\n"
        "Bot splits and sends all parts in the same format.\n\n"

        "🔗 *Merge Files*\n"
        "Upload multiple files one by one, then tap Finish Merging.\n"
        "Bot merges all into a single file.\n\n"

        "✏️ *File Editor*\n"
        "Upload a file to view, add, edit, or remove contacts.\n"
        "When done, tap Finish & Download.\n\n"

        "📋 *List Maker*\n"
        "Enter a list name, then upload WhatsApp screenshots.\n"
        "Add captions in format `GroupName | Count` for auto-extraction.\n"
        "Bot generates a formatted list with totals.\n\n"

        "📝 *Rename File*\n"
        "Upload any file and provide a new name.\n"
        "Bot sends it back with the new filename.\n\n"

        "📛 *Rename Contact*\n"
        "Upload a VCF file and enter a new base name.\n"
        "All contacts are renamed while preserving their numbering.\n\n"

        "⚙️ *Settings*\n"
        "View and edit default values used by VCF Maker.\n\n"

        "🔄 *Reset*\n"
        "Reset all settings to factory defaults.\n\n"

        "_Tip: Every screen has 🔙 Back and ❌ Cancel buttons._"
    )
    await update.effective_message.reply_text(
        text,
        reply_markup=main_menu_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return MAIN_MENU


# ─────────────────────────────────────────────────────────────────────────────
# UNKNOWN / FALLBACK
# ─────────────────────────────────────────────────────────────────────────────

async def unknown_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text(
        "⚠️ I didn't understand that. Use /start to open the menu.",
        reply_markup=ReplyKeyboardMarkup(
            [["🏠 Open Menu"]],
            resize_keyboard=True,
            one_time_keyboard=True,
        ),
    )


async def open_menu_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await cmd_start(update, context)


# ─────────────────────────────────────────────────────────────────────────────
# CANCEL FILTER
# ─────────────────────────────────────────────────────────────────────────────

CANCEL_FILTER = filters.Regex(f"^{re.escape(BTN_CANCEL)}$")
BACK_FILTER = filters.Regex(f"^{re.escape(BTN_BACK)}$")


# ─────────────────────────────────────────────────────────────────────────────
# BUILD ConversationHandler
# ─────────────────────────────────────────────────────────────────────────────

def build_conv_handler() -> ConversationHandler:
    text_filter = filters.TEXT & ~filters.COMMAND

    return ConversationHandler(
        entry_points=[CommandHandler("start", cmd_start)],
        states={
            # ── Main Menu ────────────────────────────────────────────────────
            MAIN_MENU: [
                MessageHandler(text_filter, main_menu_router),
            ],

            # ── File Analysis ────────────────────────────────────────────────
            ANALYSIS_UPLOAD: [
                MessageHandler(filters.Document.ALL, analysis_upload_handler),
                MessageHandler(text_filter, analysis_upload_handler),
            ],

            # ── File Converter ───────────────────────────────────────────────
            CONVERT_UPLOAD: [
                MessageHandler(filters.Document.ALL, converter_upload_handler),
                MessageHandler(text_filter, converter_upload_handler),
            ],
            CONVERT_FORMAT: [
                MessageHandler(text_filter, converter_format_handler),
            ],

            # ── Quick VCF ────────────────────────────────────────────────────
            QVCF_FILENAME: [
                MessageHandler(text_filter, qvcf_filename),
            ],
            QVCF_CNAME: [
                MessageHandler(text_filter, qvcf_cname),
            ],
            QVCF_CNUMBER: [
                MessageHandler(text_filter, qvcf_cnumber),
            ],
            QVCF_MORE: [
                MessageHandler(text_filter, qvcf_more_handler),
            ],

            # ── VCF Maker ────────────────────────────────────────────────────
            MAKER_FILENAME: [
                MessageHandler(text_filter, maker_filename),
            ],
            MAKER_CNAME: [
                MessageHandler(text_filter, maker_cname),
            ],
            MAKER_PER_FILE: [
                MessageHandler(text_filter, maker_per_file),
            ],
            MAKER_NUM_START: [
                MessageHandler(text_filter, maker_num_start),
            ],
            MAKER_FILE_START: [
                MessageHandler(text_filter, maker_file_start),
            ],
            MAKER_COUNTRY: [
                MessageHandler(text_filter, maker_country),
            ],
            MAKER_GROUP: [
                MessageHandler(text_filter, maker_group),
            ],
            MAKER_CONFIRM: [
                MessageHandler(text_filter, maker_confirm_handler),
            ],
            MAKER_EDIT_FIELD: [
                MessageHandler(text_filter, maker_edit_field_handler),
            ],
            MAKER_UPLOAD: [
                MessageHandler(filters.Document.ALL, maker_upload_handler),
                MessageHandler(text_filter, maker_upload_handler),
            ],
            MAKER_OUTPUT_MODE: [
                MessageHandler(text_filter, maker_output_mode_handler),
            ],

            # ── Split File ───────────────────────────────────────────────────
            SPLIT_UPLOAD: [
                MessageHandler(filters.Document.ALL, split_upload_handler),
                MessageHandler(text_filter, split_upload_handler),
            ],
            SPLIT_COUNT: [
                MessageHandler(text_filter, split_count_handler),
            ],

            # ── Merge Files ──────────────────────────────────────────────────
            MERGE_UPLOAD: [
                MessageHandler(filters.Document.ALL, merge_upload_handler),
                MessageHandler(text_filter, merge_upload_handler),
            ],

            # ── File Editor ──────────────────────────────────────────────────
            EDITOR_UPLOAD: [
                MessageHandler(filters.Document.ALL, editor_upload_handler),
                MessageHandler(text_filter, editor_upload_handler),
            ],
            EDITOR_MENU: [
                MessageHandler(text_filter, editor_menu_handler),
            ],
            EDITOR_REMOVE: [
                MessageHandler(text_filter, editor_remove_handler),
            ],
            EDITOR_EDIT_SEL: [
                MessageHandler(text_filter, editor_edit_select_handler),
            ],
            EDITOR_EDIT_NAME: [
                MessageHandler(text_filter, editor_edit_name_handler),
            ],
            EDITOR_EDIT_NUM: [
                MessageHandler(text_filter, editor_edit_num_handler),
            ],
            EDITOR_ADD_NAME: [
                MessageHandler(text_filter, editor_add_name_handler),
            ],
            EDITOR_ADD_NUM: [
                MessageHandler(text_filter, editor_add_num_handler),
            ],

            # ── List Maker ───────────────────────────────────────────────────
            LIST_NAME: [
                MessageHandler(text_filter, listmaker_name_handler),
            ],
            LIST_UPLOAD: [
                MessageHandler(filters.PHOTO, listmaker_upload_handler),
                MessageHandler(filters.Document.ALL, listmaker_upload_handler),
                MessageHandler(text_filter, listmaker_upload_handler),
            ],

            # ── Rename File ──────────────────────────────────────────────────
            RENFILE_UPLOAD: [
                MessageHandler(filters.Document.ALL, renfile_upload_handler),
                MessageHandler(text_filter, renfile_upload_handler),
            ],
            RENFILE_NAME: [
                MessageHandler(text_filter, renfile_name_handler),
            ],

            # ── Rename Contact ───────────────────────────────────────────────
            RENCONT_UPLOAD: [
                MessageHandler(filters.Document.ALL, rencont_upload_handler),
                MessageHandler(text_filter, rencont_upload_handler),
            ],
            RENCONT_NAME: [
                MessageHandler(text_filter, rencont_name_handler),
            ],

            # ── Settings ─────────────────────────────────────────────────────
            SETTINGS_MENU: [
                MessageHandler(text_filter, settings_menu_handler),
            ],
            SETTINGS_EDIT: [
                MessageHandler(text_filter, settings_edit_handler),
            ],

            # ── Reset ────────────────────────────────────────────────────────
            RESET_CONFIRM: [
                MessageHandler(text_filter, reset_confirm_handler),
            ],
        },
        fallbacks=[
            CommandHandler("start", cmd_start),
            MessageHandler(CANCEL_FILTER, cancel_handler),
        ],
        allow_reentry=True,
        per_message=False,
    )


# ─────────────────────────────────────────────────────────────────────────────
# MAIN ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    """Start the VCF Master Bot."""
    application = Application.builder().token(BOT_TOKEN).build()

    # Register the master conversation handler
    application.add_handler(build_conv_handler())

    # Fallback for messages outside a conversation
    application.add_handler(
        MessageHandler(filters.Regex(r"^🏠 Open Menu$"), open_menu_handler)
    )
    application.add_handler(
        MessageHandler(filters.ALL & ~filters.COMMAND, unknown_message)
    )

    logger.info("📱 VCF Master Bot (ReplyKeyboard) is running...")
    application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
